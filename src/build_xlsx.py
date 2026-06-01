"""
Rebuild outputs/ebola_timeseries.xlsx from data/timeseries.csv,
data/country_breakdown.csv, and data/notes.md.

The CSVs and notes.md are the source of truth. The XLSX is a build
artifact for users who want to view the tracker in Excel. It is not
required by generate_charts.py (which reads CSVs directly).

Usage:
    python3 src/build_xlsx.py
"""

from __future__ import annotations

import csv
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --- Paths -------------------------------------------------------------------

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "outputs"

TIMESERIES_CSV = DATA_DIR / "timeseries.csv"
COUNTRY_CSV = DATA_DIR / "country_breakdown.csv"
NOTES_MD = DATA_DIR / "notes.md"
XLSX_OUT = OUTPUT_DIR / "ebola_timeseries.xlsx"

# --- Style constants ---------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", size=10, color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="E8F0FE")
DATA_FONT = Font(name="Arial", size=10)
SOURCE_FONT = Font(name="Arial", size=9, italic=True, color="555555")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Arial", size=10, italic=True, color="555555")
NOTES_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="1F4E79")
NOTE_FONT = Font(name="Arial", size=10, color="333333")

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# --- Helpers -----------------------------------------------------------------

def write_headers(ws, headers, row, widths=None):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row, c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
    if widths:
        for c, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 32


def apply_data_style(ws, row, n_cols, is_alt=False, source_cols=None):
    """source_cols is a 1-indexed tuple of columns that get the gray italic font."""
    source_cols = source_cols or ()
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        if c in source_cols:
            cell.font = SOURCE_FONT
        else:
            cell.font = DATA_FONT
        if is_alt:
            cell.fill = ALT_FILL


# --- Main sheet --------------------------------------------------------------

def build_main_sheet(wb, ts_csv: Path, notes_md: Path):
    ws = wb.create_sheet("Ebola Daily Time Series")

    # Title
    ws.cell(1, 1, value="2026 Ebola (Bundibugyo) Outbreak: Daily Case Counts (global)")
    ws.cell(1, 1).font = TITLE_FONT
    ws.merge_cells("A1:H1")

    # Subtitle / description
    ws.cell(2, 1, value=(
        "Global outbreak totals (all reporting countries). Compiled from "
        "WHO AFRO Sitrep, WHO DON, CDC, Africa CDC, BNO News. "
        "Per-country detail in the Country Breakdown sheet."
    ))
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = LEFT_WRAP
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    # Headers (row 4)
    headers = [
        "Report Date",
        "Suspected Cases\n(Global)",
        "Confirmed Cases\n(Global)",
        "Total Cases\n(Global)",
        "Suspected Deaths\n(Global)",
        "Confirmed Deaths\n(Global)",
        "Primary Source",
        "Source Timestamp",
    ]
    widths = [16, 16, 16, 16, 16, 16, 48, 16]
    write_headers(ws, headers, row=4, widths=widths)

    # Data rows (starting row 5)
    data_start = 5
    with open(ts_csv, newline="") as f:
        reader = csv.DictReader(f)
        for i, r in enumerate(reader):
            row = data_start + i
            ws.cell(row, 1, value=r["report_date"])
            ws.cell(row, 2, value=_int_or_none(r["suspected_global"]))
            ws.cell(row, 3, value=_int_or_none(r["confirmed_global"]))
            ws.cell(row, 4, value=_int_or_none(r["total_global"]))
            ws.cell(row, 5, value=_int_or_none(r["suspected_deaths_global"]))
            ws.cell(row, 6, value=_int_or_none(r["confirmed_deaths_global"]))
            ws.cell(row, 7, value=r["primary_source"])
            ws.cell(row, 8, value=r["source_timestamp"])
            apply_data_style(ws, row, n_cols=8, is_alt=(i % 2 == 1),
                             source_cols=(7, 8))
        last_data_row = data_start + i

    # NOTES block
    notes_row = last_data_row + 2
    ws.cell(notes_row, 1, value="NOTES")
    ws.cell(notes_row, 1).font = NOTES_HEADER_FONT
    notes_row += 1
    with open(notes_md) as f:
        for line in f:
            line = line.rstrip()
            if not line.startswith("- "):
                continue
            bullet_text = "• " + line[2:]
            ws.cell(notes_row, 1, value=bullet_text)
            ws.cell(notes_row, 1).font = NOTE_FONT
            ws.cell(notes_row, 1).alignment = LEFT_WRAP
            ws.merge_cells(start_row=notes_row, start_column=1,
                           end_row=notes_row, end_column=8)
            notes_row += 1

    ws.freeze_panes = "A5"
    return ws


# --- Charts sheet (derived) --------------------------------------------------

def build_charts_sheet(wb, ts_csv: Path, window: int = 7):
    ws = wb.create_sheet("Charts")

    ws.cell(1, 1, value="7-day rolling sums of daily changes (global)")
    ws.cell(1, 1).font = TITLE_FONT
    ws.merge_cells("A1:J1")

    ws.cell(2, 1, value=(
        "Derived from the main timeseries. Charts live as PNG files in "
        "outputs/ — this table is the data behind them."
    ))
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = LEFT_WRAP
    ws.merge_cells("A2:J2")

    headers = [
        "Date",
        "Daily Δ Suspected",
        "Daily Δ Confirmed",
        "Daily Δ Suspected Deaths",
        "Daily Δ Confirmed Deaths",
        "Suspected cases (7-day sum)",
        "Confirmed cases (7-day sum)",
        "Suspected deaths (7-day sum)",
        "Confirmed deaths (7-day sum)",
        "Window (days)",
    ]
    widths = [14, 18, 18, 22, 22, 24, 24, 24, 24, 14]
    write_headers(ws, headers, row=4, widths=widths)

    # Compute deltas + rolling sums from CSV
    rows = []
    with open(ts_csv, newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            rows.append({
                "date": r["report_date"],
                "sus": int(r["suspected_global"] or 0),
                "conf": int(r["confirmed_global"] or 0),
                "deaths": int(r["suspected_deaths_global"] or 0),
                "conf_deaths": int(r["confirmed_deaths_global"] or 0),
            })

    prev = {"sus": 0, "conf": 0, "deaths": 0, "conf_deaths": 0}
    deltas = []
    for cur in rows:
        deltas.append({
            "date": cur["date"],
            "d_sus": cur["sus"] - prev["sus"],
            "d_conf": cur["conf"] - prev["conf"],
            "d_deaths": cur["deaths"] - prev["deaths"],
            "d_conf_deaths": cur["conf_deaths"] - prev["conf_deaths"],
        })
        prev = cur

    for i, row in enumerate(deltas):
        win = deltas[max(0, i - (window - 1)): i + 1]
        row["r7_sus"] = sum(w["d_sus"] for w in win)
        row["r7_conf"] = sum(w["d_conf"] for w in win)
        row["r7_deaths"] = sum(w["d_deaths"] for w in win)
        row["r7_conf_deaths"] = sum(w["d_conf_deaths"] for w in win)
        row["window"] = len(win)

        r = 5 + i
        ws.cell(r, 1, value=row["date"])
        ws.cell(r, 2, value=row["d_sus"])
        ws.cell(r, 3, value=row["d_conf"])
        ws.cell(r, 4, value=row["d_deaths"])
        ws.cell(r, 5, value=row["d_conf_deaths"])
        ws.cell(r, 6, value=row["r7_sus"])
        ws.cell(r, 7, value=row["r7_conf"])
        ws.cell(r, 8, value=row["r7_deaths"])
        ws.cell(r, 9, value=row["r7_conf_deaths"])
        ws.cell(r, 10, value=row["window"])
        apply_data_style(ws, r, n_cols=10, is_alt=(i % 2 == 1))

    return ws


# --- Country breakdown sheet -------------------------------------------------

def build_country_sheet(wb, country_csv: Path):
    ws = wb.create_sheet("Country Breakdown")

    ws.cell(1, 1, value="2026 Ebola Outbreak: Country Breakdown (long format)")
    ws.cell(1, 1).font = TITLE_FONT
    ws.merge_cells("A1:H1")

    ws.cell(2, 1, value=(
        "Per-country case and death counts derived from the main timeseries. "
        "For dates where only a global aggregate has been published, DRC values "
        "are derived as (Global - Uganda). Add new countries here as they begin "
        "reporting; the main sheet then accumulates the new totals into its "
        "Global columns."
    ))
    ws.cell(2, 1).font = SUBTITLE_FONT
    ws.cell(2, 1).alignment = LEFT_WRAP
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 45

    headers = [
        "Date",
        "Country",
        "Suspected",
        "Confirmed",
        "Suspected Deaths",
        "Confirmed Deaths",
        "Primary Source",
        "Source Timestamp",
    ]
    widths = [14, 12, 14, 14, 18, 18, 48, 16]
    write_headers(ws, headers, row=4, widths=widths)

    with open(country_csv, newline="") as f:
        reader = csv.DictReader(f)
        for i, r in enumerate(reader):
            row = 5 + i
            ws.cell(row, 1, value=r["date"])
            ws.cell(row, 2, value=r["country"])
            ws.cell(row, 3, value=_int_or_none(r["suspected"]))
            ws.cell(row, 4, value=_int_or_none(r["confirmed"]))
            ws.cell(row, 5, value=_int_or_none(r["suspected_deaths"]))
            ws.cell(row, 6, value=_int_or_none(r["confirmed_deaths"]))
            ws.cell(row, 7, value=r["primary_source"])
            ws.cell(row, 8, value=r["source_timestamp"])
            apply_data_style(ws, row, n_cols=8, is_alt=(i % 2 == 1),
                             source_cols=(7, 8))

    return ws


# --- Helpers -----------------------------------------------------------------

def _int_or_none(s):
    if s is None or s == "":
        return None
    try:
        return int(s)
    except ValueError:
        return s


# --- Entry point -------------------------------------------------------------

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Workbook() creates a default sheet; remove it so our named sheets
    # come out in the order we want.
    wb.remove(wb.active)

    build_main_sheet(wb, TIMESERIES_CSV, NOTES_MD)
    build_charts_sheet(wb, TIMESERIES_CSV)
    build_country_sheet(wb, COUNTRY_CSV)

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
